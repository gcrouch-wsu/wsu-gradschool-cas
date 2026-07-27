from __future__ import annotations

import json
import os
import subprocess
import sys
import threading
import hashlib
from datetime import datetime, timezone
from html import escape
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from flask import Flask, redirect, request, url_for
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dashboard still works without row counts.
    load_workbook = None


REPO_ROOT = Path(__file__).resolve().parents[2]
BRANDING_ROOT = REPO_ROOT / ".branding-data"
UPLOAD_ROOT = BRANDING_ROOT / "uploads"
CAPTURE_MANIFEST_FILE = BRANDING_ROOT / "capture-manifest.json"
PROFILES = {
    "gradcas": {
        "label": "GradCAS",
        "xlsx": REPO_ROOT / "GradCAS.xlsx",
    },
    "engineeringcas": {
        "label": "EngineeringCAS",
        "xlsx": REPO_ROOT / "EngCAS.xlsx",
    },
}

app = Flask(__name__)
jobs: dict[str, dict[str, Any]] = {}


def selected_profile() -> str:
    profile = (request.args.get("profile") or "gradcas").strip().lower()
    return profile if profile in PROFILES else "gradcas"


def profile_redirect(profile: str):
    return redirect(url_for("index", profile=profile))


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def load_env_file(target: dict[str, str], path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'\"")
        if value or not target.get(key):
            target[key] = value


def process_env() -> dict[str, str]:
    env = os.environ.copy()
    load_env_file(env, REPO_ROOT / ".env.local")
    load_env_file(env, REPO_ROOT / ".env.branding")
    return env


def branding_start_url(env: dict[str, str]) -> str:
    return env.get("BRANDING_START_URL", "https://webadmit.org/").strip()


def blob_token_source(env: dict[str, str]) -> str:
    if env.get("BLOB_READ_WRITE_TOKEN"):
        if (REPO_ROOT / ".env.local").exists():
            local_text = (REPO_ROOT / ".env.local").read_text(encoding="utf-8", errors="ignore")
            if "BLOB_READ_WRITE_TOKEN=" in local_text:
                return ".env.local"
        return "environment"
    if (REPO_ROOT / ".env.local").exists():
        local_text = (REPO_ROOT / ".env.local").read_text(encoding="utf-8", errors="ignore")
        for raw_line in local_text.splitlines():
            if raw_line.strip().startswith("BLOB_READ_WRITE_TOKEN="):
                return "empty in .env.local"
    return "missing"


def save_env_value(path: Path, key: str, value: str) -> None:
    lines = []
    found = False
    if path.exists():
        lines = path.read_text(encoding="utf-8").splitlines()
    next_lines = []
    for line in lines:
        if line.strip().startswith(f"{key}="):
            next_lines.append(f"{key}={value}")
            found = True
        else:
            next_lines.append(line)
    if not found:
        next_lines.append(f"{key}={value}")
    path.write_text("\n".join(next_lines).rstrip() + "\n", encoding="utf-8")


def profile_root(profile: str) -> Path:
    return BRANDING_ROOT / "profiles" / profile


def snapshot_root(snapshot_id: str, profile: str) -> Path:
    return BRANDING_ROOT / "snapshots" / snapshot_id / profile


def status_path(profile: str) -> Path:
    return profile_root(profile) / "status.json"


def log_path(profile: str) -> Path:
    return profile_root(profile) / "flask-command.log"


def capture_state_path(profile: str) -> Path:
    return profile_root(profile) / "capture-state.json"


def read_json(path: Path) -> Any | None:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def read_text(path: Path, max_chars: int = 5000) -> str:
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""
    return text[-max_chars:]


def url_origin(value: str) -> str:
    try:
        parsed = urlparse(value)
    except Exception:
        return ""
    if not parsed.scheme or not parsed.netloc:
        return ""
    return f"{parsed.scheme.lower()}://{parsed.netloc.lower()}"


def capture_manifest() -> dict[str, Any] | None:
    value = read_json(CAPTURE_MANIFEST_FILE)
    return value if isinstance(value, dict) else None


def manifest_profile(profile: str) -> dict[str, Any] | None:
    manifest = capture_manifest()
    profiles = manifest.get("profiles") if manifest else None
    row = profiles.get(profile) if isinstance(profiles, dict) else None
    return row if isinstance(row, dict) else None


def manifest_program_ids(profile: str) -> list[str]:
    row = manifest_profile(profile)
    ids = row.get("programIds") if row else None
    if not isinstance(ids, list):
        return []
    return [str(value).strip() for value in ids if str(value).strip()]


def auth_file(profile: str) -> Path:
    return profile_root(profile) / "user.json"


def trail_file(profile: str) -> Path:
    return profile_root(profile) / "trail.json"


def capture_route_status(profile: str, env: dict[str, str]) -> dict[str, Any]:
    trail = read_json(trail_file(profile))
    has_auth = auth_file(profile).exists()
    has_trail = isinstance(trail, dict)
    has_template = bool(
        trail.get("brandingUrlTemplate") or trail.get("programUrlTemplate")
    ) if has_trail else False
    start_origin = url_origin(branding_start_url(env))
    trail_origin = url_origin(str(trail.get("loginUrl", ""))) if has_trail else ""
    same_environment = not trail_origin or not start_origin or trail_origin == start_origin
    ready = has_auth and has_template and same_environment
    if ready:
        message = "Ready. Guided login saved a Branding page route for this URL."
    elif not has_auth:
        message = "Not ready. Run guided login first."
    elif not has_trail:
        message = "Not ready. Run guided login so the app can save the Branding route."
    elif not same_environment:
        message = "Not ready. The saved guide route belongs to a different start URL."
    else:
        message = "Not ready. Guided login saved your login, but not the Branding page route."
    return {
        "ready": ready,
        "message": message,
        "hasAuth": has_auth,
        "hasTrail": has_trail,
        "hasTemplate": has_template,
        "sameEnvironment": same_environment,
    }


def write_manifest_id_file(profile: str, program_ids: list[str]) -> Path:
    jobs_root = BRANDING_ROOT / "jobs"
    jobs_root.mkdir(parents=True, exist_ok=True)
    target = jobs_root / f"{profile}-manifest-program-ids.txt"
    target.write_text("\n".join(program_ids) + "\n", encoding="utf-8")
    return target


def write_job(profile: str, **updates: Any) -> None:
    existing = jobs.get(profile, {})
    existing.update(updates)
    existing["updatedAt"] = utc_now()
    jobs[profile] = existing


def write_profile_status(profile: str, **updates: Any) -> None:
    path = status_path(profile)
    existing = read_json(path)
    if not isinstance(existing, dict):
        existing = {"profile": profile}
    existing.update(updates)
    existing["profile"] = profile
    existing["updatedAt"] = utc_now()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(existing, indent=2), encoding="utf-8")


def write_capture_state(profile: str, **updates: Any) -> None:
    path = capture_state_path(profile)
    existing = read_json(path)
    if not isinstance(existing, dict):
        existing = {}
    existing.update(updates)
    existing["profile"] = profile
    existing["updatedAt"] = utc_now()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(existing, indent=2), encoding="utf-8")


def capture_source_key(profile: str, xlsx_path: Path, manifest_ids: list[str]) -> str:
    if manifest_ids:
        digest = hashlib.sha256("\n".join(manifest_ids).encode("utf-8")).hexdigest()[:16]
        return f"manifest:{profile}:{len(manifest_ids)}:{digest}"
    try:
        stat = xlsx_path.stat()
        return f"xlsx:{xlsx_path.resolve()}:{int(stat.st_mtime)}:{stat.st_size}"
    except OSError:
        return f"xlsx:{xlsx_path}"


def resumable_snapshot_id(profile: str, source_key: str) -> str | None:
    state = read_json(capture_state_path(profile))
    if not isinstance(state, dict):
        return None
    if state.get("status") == "completed":
        return None
    if state.get("sourceKey") != source_key:
        return None
    snapshot_id = str(state.get("snapshotId") or "")
    if not snapshot_id:
        return None
    if not snapshot_root(snapshot_id, profile).exists():
        return None
    return snapshot_id


def latest_local_manifest(profile: str) -> dict[str, Any] | None:
    snapshots = BRANDING_ROOT / "snapshots"
    if not snapshots.exists():
        return None
    best: dict[str, Any] | None = None
    for snapshot_dir in snapshots.iterdir():
        manifest = read_json(snapshot_dir / profile / "manifest.json")
        if not manifest or manifest.get("status") != "completed":
            continue
        if not best or str(manifest.get("createdAt", "")) > str(best.get("createdAt", "")):
            best = manifest
    return best


def latest_upload(profile: str) -> Path | None:
    upload_dir = UPLOAD_ROOT / profile
    if not upload_dir.exists():
        return None
    files = [path for path in upload_dir.iterdir() if path.suffix.lower() in {".xlsx", ".xls"}]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def count_program_ids(xlsx_path: Path) -> int | None:
    if load_workbook is None or not xlsx_path.exists() or xlsx_path.suffix.lower() != ".xlsx":
        return None
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = workbook["Program Attributes"] if "Program Attributes" in workbook.sheetnames else workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return None
        program_id_index = None
        for index, value in enumerate(headers):
            if str(value or "").strip().lower() in {"program id", "programid"}:
                program_id_index = index
                break
        if program_id_index is None:
            return None
        count = 0
        seen: set[str] = set()
        for row in rows:
            if program_id_index >= len(row):
                continue
            value = str(row[program_id_index] or "").strip()
            if value and value not in seen:
                seen.add(value)
                count += 1
        workbook.close()
        return count
    except Exception:
        return None


def progress_line(status: dict[str, Any], program_count: int | None) -> str:
    completed = status.get("completedPrograms")
    total = status.get("totalPrograms") or program_count
    if completed is not None and total:
        return f"{completed}/{total} Program IDs"
    if total:
        return f"0/{total} Program IDs ready"
    return "Program count unavailable until an .xlsx export is selected"


def run_command(profile: str, label: str, command: list[str]) -> None:
    profile_root(profile).mkdir(parents=True, exist_ok=True)
    log_file = log_path(profile)
    log_file.write_text(f"[{utc_now()}] Running: {' '.join(command)}\n\n", encoding="utf-8")
    write_job(
        profile,
        action=label,
        status="running",
        startedAt=utc_now(),
        message=f"Running. Log: {log_file}",
    )
    try:
        result = subprocess.run(
            command,
            cwd=REPO_ROOT,
            env=process_env(),
            text=True,
            capture_output=True,
            check=False,
        )
        with log_file.open("a", encoding="utf-8") as handle:
            if result.stdout:
                handle.write("\n[stdout]\n")
                handle.write(result.stdout)
            if result.stderr:
                handle.write("\n[stderr]\n")
                handle.write(result.stderr)
        if result.returncode != 0:
            write_job(
                profile,
                status="error",
                completedAt=utc_now(),
                message=result.stderr.strip() or result.stdout.strip() or "Command failed",
            )
            return
        if label == "upload_latest":
            message = "Branding upload complete. You can close this dashboard window."
            write_profile_status(
                profile,
                mode="upload",
                status="completed",
                completedAt=utc_now(),
                message=message,
            )
        else:
            message = result.stdout.strip() or "Completed"
        write_job(
            profile,
            status="completed",
            completedAt=utc_now(),
            message=message,
        )
    except Exception as exc:
        write_job(profile, status="error", completedAt=utc_now(), message=str(exc))


def run_popen_command(profile: str, label: str, command: list[str]) -> None:
    profile_root(profile).mkdir(parents=True, exist_ok=True)
    log_file = log_path(profile)
    write_job(
        profile,
        action=label,
        status="running",
        startedAt=utc_now(),
        message=f"Starting. Log: {log_file}",
    )
    try:
        with log_file.open("w", encoding="utf-8") as handle:
            handle.write(f"[{utc_now()}] Starting: {' '.join(command)}\n\n")
            handle.flush()
            process = subprocess.Popen(
                command,
                cwd=REPO_ROOT,
                env=process_env(),
                text=True,
                stdout=handle,
                stderr=subprocess.STDOUT,
                stdin=subprocess.DEVNULL,
            )
            write_job(
                profile,
                pid=process.pid,
                message=f"Started process {process.pid}. Close the browser window when finished.",
            )
            return_code = process.wait()
            handle.write(f"\n[{utc_now()}] Process exited with code {return_code}\n")
        if return_code == 0:
            write_job(
                profile,
                status="completed",
                completedAt=utc_now(),
                message="Guided login completed. Auth and navigation trail were saved.",
            )
        else:
            write_job(
                profile,
                status="error",
                completedAt=utc_now(),
                message=f"Guided login failed with exit code {return_code}. See command log.",
            )
    except Exception as exc:
        write_job(profile, status="error", completedAt=utc_now(), message=str(exc))


def upload_latest_snapshot(profile: str) -> None:
    manifest = latest_local_manifest(profile)
    if not manifest:
        write_job(
            profile,
            action="upload_latest",
            status="error",
            completedAt=utc_now(),
            message="No completed local snapshot found for upload.",
        )
        return
    snapshot_id = str(manifest.get("snapshotId") or "")
    command = [
        "node",
        "tools/branding/upload-snapshot.mjs",
        "--profile",
        profile,
        "--snapshot-id",
        snapshot_id,
    ]
    run_command(profile, "upload_latest", command)


def load_capture_manifest(profile: str) -> None:
    profile_root(profile).mkdir(parents=True, exist_ok=True)
    log_file = log_path(profile)
    command = ["node", "tools/branding/read-capture-manifest.mjs"]
    log_file.write_text(f"[{utc_now()}] Loading capture manifest: {' '.join(command)}\n\n", encoding="utf-8")
    write_job(
        profile,
        action="load_manifest",
        status="running",
        startedAt=utc_now(),
        message="Loading capture manifest from Vercel Blob.",
    )
    result = subprocess.run(
        command,
        cwd=REPO_ROOT,
        env=process_env(),
        text=True,
        capture_output=True,
        check=False,
    )
    with log_file.open("a", encoding="utf-8") as handle:
        if result.stdout:
            handle.write("\n[stdout]\n")
            handle.write(result.stdout)
        if result.stderr:
            handle.write("\n[stderr]\n")
            handle.write(result.stderr)
    if result.returncode != 0:
        write_job(
            profile,
            status="error",
            completedAt=utc_now(),
            message=result.stderr.strip() or result.stdout.strip() or "Could not load capture manifest.",
        )
        return
    try:
        parsed = json.loads(result.stdout)
    except Exception as exc:
        write_job(profile, status="error", completedAt=utc_now(), message=f"Invalid manifest JSON: {exc}")
        return
    CAPTURE_MANIFEST_FILE.parent.mkdir(parents=True, exist_ok=True)
    CAPTURE_MANIFEST_FILE.write_text(json.dumps(parsed, indent=2), encoding="utf-8")
    write_job(
        profile,
        status="completed",
        completedAt=utc_now(),
        message=f"Loaded capture manifest for {parsed.get('publicationTitle', 'current publication')}.",
    )


def start_thread(profile: str, label: str, command: list[str]) -> None:
    target = run_popen_command if label == "guide" else run_command
    thread = threading.Thread(target=target, args=(profile, label, command), daemon=True)
    thread.start()


def capture_and_upload(profile: str, xlsx_path: Path, delay_ms: int) -> None:
    env = process_env()
    start_url = branding_start_url(env)
    manifest_ids = manifest_program_ids(profile)
    source_key = capture_source_key(profile, xlsx_path, manifest_ids)
    resume_snapshot_id = resumable_snapshot_id(profile, source_key)
    snapshot_id = resume_snapshot_id or datetime.now(timezone.utc).isoformat().replace(":", "-").replace(".", "-")
    output_dir = snapshot_root(snapshot_id, profile)
    status_file = status_path(profile)
    auth_file = profile_root(profile) / "user.json"
    trail_file = profile_root(profile) / "trail.json"
    id_file = write_manifest_id_file(profile, manifest_ids) if manifest_ids else None
    export_command = [
        "node",
        "tools/branding/cli.mjs",
        "export",
        "--profile",
        profile,
        "--output-dir",
        str(output_dir),
        "--auth-file",
        str(auth_file),
        "--trail-file",
        str(trail_file),
        "--status-file",
        str(status_file),
        "--login-url",
        start_url,
        "--delay-ms",
        str(delay_ms),
        "--non-interactive",
    ]
    if id_file:
        export_command.extend(["--id-file", str(id_file)])
    else:
        export_command.extend(["--xlsx", str(xlsx_path)])
    upload_command = [
        "node",
        "tools/branding/upload-snapshot.mjs",
        "--profile",
        profile,
        "--snapshot-id",
        snapshot_id,
    ]

    write_job(
        profile,
        action="capture",
        status="running",
        startedAt=utc_now(),
        message=(
            f"Resuming capture/upload for snapshot {snapshot_id}"
            if resume_snapshot_id
            else f"Capturing {len(manifest_ids)} Program IDs from Vercel manifest"
            if manifest_ids
            else f"Capturing {xlsx_path.name}"
        ),
        snapshotId=snapshot_id,
    )
    write_capture_state(
        profile,
        status="capturing",
        snapshotId=snapshot_id,
        sourceKey=source_key,
        outputDir=str(output_dir),
        totalPrograms=len(manifest_ids) if manifest_ids else count_program_ids(xlsx_path),
        resumed=bool(resume_snapshot_id),
    )
    result = subprocess.run(
        export_command,
        cwd=REPO_ROOT,
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        write_job(
            profile,
            status="error",
            completedAt=utc_now(),
            message=result.stderr.strip() or result.stdout.strip() or "Capture failed",
        )
        write_capture_state(profile, status="capture_error", snapshotId=snapshot_id)
        return

    write_job(profile, status="running", message="Capture completed. Uploading to Vercel Blob. If interrupted, upload will resume from saved state.")
    write_profile_status(
        profile,
        mode="upload",
        status="running",
        snapshotId=snapshot_id,
        message="Capture completed. Uploading to Vercel Blob. If interrupted, upload will resume from saved state.",
    )
    write_capture_state(profile, status="uploading", snapshotId=snapshot_id)
    upload = subprocess.run(
        upload_command,
        cwd=REPO_ROOT,
        env=env,
        text=True,
        capture_output=True,
        check=False,
    )
    if upload.returncode != 0:
        write_job(
            profile,
            status="error",
            completedAt=utc_now(),
            message=upload.stderr.strip() or upload.stdout.strip() or "Upload failed",
        )
        write_profile_status(
            profile,
            mode="upload",
            status="error",
            snapshotId=snapshot_id,
            completedAt=utc_now(),
            message="Upload was interrupted or failed. Start capture again to resume from saved state.",
        )
        write_capture_state(profile, status="upload_error", snapshotId=snapshot_id)
        return

    completion_message = "Branding upload complete. You can close this dashboard window."
    write_job(
        profile,
        status="completed",
        completedAt=utc_now(),
        message=completion_message,
        snapshotId=snapshot_id,
    )
    write_profile_status(
        profile,
        mode="upload",
        status="completed",
        snapshotId=snapshot_id,
        completedAt=utc_now(),
        message=completion_message,
    )
    write_capture_state(
        profile,
        status="completed",
        snapshotId=snapshot_id,
        completedAt=utc_now(),
    )


def page_shell(body: str) -> str:
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta http-equiv="refresh" content="5">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>CAS Branding Capture</title>
  <style>
    :root {{
      --bg: #f4efe6;
      --ink: #211b15;
      --muted: #6f6256;
      --card: #fffaf2;
      --line: #d9cbbb;
      --accent: #9b2f24;
      --accent-2: #315f52;
      --shadow: 0 22px 60px rgba(62, 39, 20, .16);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      color: var(--ink);
      font-family: Georgia, "Times New Roman", serif;
      background:
        radial-gradient(circle at 12% 12%, rgba(155,47,36,.18), transparent 34rem),
        radial-gradient(circle at 88% 8%, rgba(49,95,82,.16), transparent 28rem),
        linear-gradient(135deg, #f8f0e5, var(--bg));
      min-height: 100vh;
    }}
    main {{ width: min(1120px, calc(100% - 32px)); margin: 0 auto; padding: 42px 0; }}
    header {{ display: flex; justify-content: space-between; gap: 24px; align-items: end; margin-bottom: 28px; }}
    h1 {{ font-size: clamp(2.4rem, 6vw, 5.5rem); line-height: .88; margin: 0; letter-spacing: -.06em; }}
    .lede {{ max-width: 520px; color: var(--muted); font: 17px/1.5 Verdana, sans-serif; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(330px, 1fr)); gap: 18px; }}
    .card {{
      background: color-mix(in srgb, var(--card) 92%, white);
      border: 1px solid var(--line);
      border-radius: 28px;
      box-shadow: var(--shadow);
      padding: 24px;
    }}
    .card h2 {{ margin: 0 0 12px; font-size: 2rem; letter-spacing: -.04em; }}
    .meta {{ display: grid; gap: 8px; margin: 16px 0; font: 14px/1.45 Verdana, sans-serif; }}
    .pill {{ display: inline-flex; width: fit-content; border: 1px solid var(--line); border-radius: 999px; padding: 6px 10px; background: white; }}
    label {{ display: block; font: 700 12px/1.3 Verdana, sans-serif; text-transform: uppercase; color: var(--muted); margin: 14px 0 6px; }}
    input {{ width: 100%; border: 1px solid var(--line); border-radius: 14px; padding: 11px 12px; background: #fffdf8; color: var(--ink); }}
    input[type="file"].sr-only {{ position: absolute; width: 1px; height: 1px; padding: 0; margin: -1px; overflow: hidden; clip: rect(0,0,0,0); white-space: nowrap; border: 0; }}
    .actions {{ display: flex; flex-wrap: wrap; gap: 10px; margin-top: 16px; }}
    button, .button {{
      border: 0;
      border-radius: 999px;
      padding: 11px 16px;
      background: var(--accent);
      color: white;
      cursor: pointer;
      font: 700 14px/1 Verdana, sans-serif;
      text-decoration: none;
    }}
    button.secondary {{ background: var(--accent-2); }}
    button:disabled {{ background: #9b938b; cursor: not-allowed; }}
    .copy-row {{ display: flex; justify-content: space-between; align-items: center; gap: 12px; margin-top: 18px; }}
    .copy-row h3 {{ margin: 0; }}
    .copy-button {{ padding: 8px 11px; background: #6f6256; font-size: 12px; }}
    .profile-tabs {{ display: flex; flex-wrap: wrap; gap: 10px; margin: 0 0 18px; }}
    .profile-tabs a {{ border: 1px solid var(--line); border-radius: 999px; padding: 10px 14px; background: white; color: var(--ink); font: 700 14px/1 Verdana, sans-serif; text-decoration: none; }}
    .profile-tabs a.active {{ background: var(--accent); border-color: var(--accent); color: white; }}
    .file-picker {{ display: inline-flex; align-items: center; justify-content: center; width: fit-content; border-radius: 999px; background: var(--accent-2); color: white; padding: 11px 16px; cursor: pointer; font: 700 14px/1 Verdana, sans-serif; }}
    .selected-file {{ margin-top: 10px; border: 1px solid #b2292e; border-radius: 14px; background: #fff6f6; padding: 11px 12px; color: var(--ink); font: 13px/1.45 Verdana, sans-serif; overflow-wrap: anywhere; }}
    .step-card {{ margin-top: 18px; border-top: 1px solid var(--line); padding-top: 18px; }}
    .step-card h3 {{ margin: 0 0 8px; font-size: 1.15rem; }}
    .steps {{ margin: 0; padding-left: 1.2rem; color: var(--muted); font: 15px/1.55 Verdana, sans-serif; }}
    .steps li {{ margin: 5px 0; }}
    pre {{ white-space: pre-wrap; overflow-wrap: anywhere; background: #211b15; color: #fff3df; padding: 14px; border-radius: 16px; font-size: 12px; max-height: 260px; overflow: auto; }}
    .warn {{ border-left: 5px solid var(--accent); padding-left: 12px; color: var(--muted); font: 14px/1.5 Verdana, sans-serif; }}
    .success {{ border-left: 5px solid var(--accent-2); background: #eef7f2; padding: 12px; color: var(--ink); font: 700 14px/1.5 Verdana, sans-serif; }}
  </style>
  <script>
    function copyText(id, button) {{
      const node = document.getElementById(id);
      if (!node) return;
      const text = node.innerText || node.textContent || "";
      const done = () => {{
        if (!button) return;
        const old = button.innerText;
        button.innerText = "Copied";
        setTimeout(() => button.innerText = old, 1200);
      }};
      if (navigator.clipboard && window.isSecureContext) {{
        navigator.clipboard.writeText(text).then(done).catch(() => fallbackCopy(text, done));
      }} else {{
        fallbackCopy(text, done);
      }}
    }}
    function fallbackCopy(text, done) {{
      const area = document.createElement("textarea");
      area.value = text;
      area.setAttribute("readonly", "");
      area.style.position = "fixed";
      area.style.left = "-9999px";
      document.body.appendChild(area);
      area.select();
      try {{ document.execCommand("copy"); done(); }} catch (_e) {{}}
      document.body.removeChild(area);
    }}
    function showChosenFile(inputId, outputId) {{
      const input = document.getElementById(inputId);
      const output = document.getElementById(outputId);
      if (!input || !output) return;
      const file = input.files && input.files[0];
      output.innerText = file ? `Selected file: ${{file.name}} (${{Math.round(file.size / 1024)}} KB). Click Save selected Excel report to store it for capture.` : output.dataset.current || "No file selected.";
    }}
  </script>
</head>
<body>
  <main>{body}</main>
</body>
</html>"""


def render_profile(profile: str, config: dict[str, Any], env: dict[str, str]) -> str:
    status = read_json(status_path(profile)) or {"status": "idle"}
    capture_state = read_json(capture_state_path(profile)) or {}
    manifest = latest_local_manifest(profile)
    job = jobs.get(profile, {})
    log_text = read_text(log_path(profile))
    token_present = bool(env.get("BLOB_READ_WRITE_TOKEN"))
    uploaded_xlsx = latest_upload(profile)
    xlsx = uploaded_xlsx or config["xlsx"]
    capture_row = manifest_profile(profile)
    manifest_ids = manifest_program_ids(profile)
    program_count = len(manifest_ids) if manifest_ids else count_program_ids(Path(xlsx))
    progress = progress_line(status, program_count)
    route_status = capture_route_status(profile, env)
    can_capture = token_present and route_status["ready"] and bool(program_count)
    status_id = f"{profile}-status"
    log_id = f"{profile}-log"
    file_output_id = f"{profile}-file-output"
    file_input_id = f"{profile}-file"
    manifest_source_line = "No Vercel capture manifest loaded. Capture will use the selected Excel report."
    if capture_row:
        latest_snapshot = capture_row.get("latestSnapshotId") or "none"
        manifest_source_line = (
            f"Vercel manifest: {len(manifest_ids)} Program IDs for "
            f"{escape(str(capture_row.get('label') or config['label']))}; latest snapshot {escape(str(latest_snapshot))}."
        )
    manifest_line = "No completed local capture yet."
    if manifest:
        manifest_line = (
            f"Snapshot {escape(str(manifest.get('snapshotId', '')))}; "
            f"ok {escape(str(manifest.get('okPrograms', 0)))}; "
            f"empty {escape(str(manifest.get('emptyShellPrograms', 0)))}; "
            f"errors {escape(str(manifest.get('errorPrograms', 0)))}"
        )
    program_source_note = (
        f"Program IDs are loaded from the Vercel manifest for {escape(config['label'])}. "
        "Skip the Excel buttons unless you want to use a report as a fallback."
        if capture_row
        else "No Vercel manifest is loaded for this profile. Choose and save an Excel report so capture has Program IDs."
    )
    excel_details_attr = "" if capture_row else " open"
    capture_disabled = "" if can_capture else " disabled"
    capture_ready_note = (
        f"Ready to capture {program_count} Program IDs."
        if can_capture
        else "Capture unlocks when the Blob token, Program IDs, and guided-login Branding route are ready."
    )
    resumable_note = ""
    if isinstance(capture_state, dict) and capture_state.get("status") not in {None, "completed"}:
        resumable_note = (
            f"Saved state found for snapshot {escape(str(capture_state.get('snapshotId', '')))}. "
            "Starting capture again will resume that snapshot and skip completed Program IDs."
        )
    completion_note = ""
    if status.get("mode") == "upload" and status.get("status") == "completed":
        completion_note = "Branding upload complete. You can close this dashboard window."
    return f"""
    <section class="card">
      <h2>{escape(config["label"])} branding capture</h2>
      <div class="meta">
        <span class="pill">Collector: {escape(str(status.get("status", "idle")))}</span>
        <span class="pill">Blob token: {"present" if token_present else "missing"}</span>
        <span class="pill">Branding route: {"ready" if route_status["ready"] else "not ready"}</span>
        <span class="pill">Progress: {escape(progress)}</span>
        <span><strong>Current Excel report:</strong> {escape(str(xlsx))}</span>
        <span><strong>Capture source:</strong> {manifest_source_line}</span>
        <span><strong>Guided login:</strong> {escape(str(route_status["message"]))}</span>
        <span>Latest local: {manifest_line}</span>
      </div>
      {f'<p class="success">{completion_note}</p>' if completion_note else ''}
      <div class="step-card">
        <h3>1. Program IDs</h3>
        <p class="warn">{program_source_note}</p>
        <details{excel_details_attr}>
          <summary><strong>Excel fallback</strong></summary>
          <form method="post" action="{url_for('upload_export', profile=profile)}" enctype="multipart/form-data">
            <label for="{file_input_id}">Excel export</label>
            <label class="file-picker" for="{file_input_id}">Choose Excel report</label>
            <input class="sr-only" id="{file_input_id}" name="xlsx_file" type="file" accept=".xlsx,.xls" onchange="showChosenFile('{file_input_id}', '{file_output_id}')">
            <div id="{file_output_id}" class="selected-file" data-current="Current saved report: {escape(str(xlsx))}">Current saved report: {escape(str(xlsx))}</div>
            <div class="actions">
              <button type="submit" class="secondary">Save selected Excel report</button>
            </div>
          </form>
        </details>
      </div>
      <div class="step-card">
        <h3>2. Save the Branding route</h3>
        <p class="warn">Edge will open. Log into WebAdMIT, go to CAS Configuration Portal, choose {escape(config["label"])} and the live cycle, click Details for a program, then click Branding. Wait until the Branding page is visible, repeat for 2 or 3 programs if possible, and close Edge while still on a Branding page. Closing the window saves the login, cycle route, and Program ID URL pattern.</p>
        <form method="post" action="{url_for('guide', profile=profile)}">
          <div class="actions">
            <button type="submit" class="secondary">Open guided login</button>
          </div>
        </form>
      </div>
      <div class="step-card">
        <h3>3. Capture and upload</h3>
        <p class="warn">{capture_ready_note} Capture opens Edge, visits each Program ID, writes the branding snapshot, and uploads it automatically. If it is interrupted, press this same button again to resume from saved state. When the dashboard says "Branding upload complete," you can close this dashboard window.</p>
        {f'<p class="warn">{resumable_note}</p>' if resumable_note else ''}
        <form method="post" action="{url_for('capture', profile=profile)}">
          <input type="hidden" name="xlsx" value="{escape(str(xlsx))}">
          <label for="{profile}-delay">Delay per Program ID, ms</label>
          <input id="{profile}-delay" name="delay_ms" value="4500">
          <div class="actions">
            <button type="submit"{capture_disabled}>Start {escape(config["label"])} capture and upload</button>
          </div>
        </form>
      </div>
      <details class="step-card">
        <summary><strong>Fallback: upload latest completed snapshot</strong></summary>
        <p class="warn">Use this only if capture completed locally but the automatic upload did not finish. It does not recapture pages.</p>
        <form method="post" action="{url_for('upload_latest', profile=profile)}">
          <div class="actions">
            <button type="submit" class="secondary">Upload latest completed snapshot</button>
          </div>
        </form>
      </details>
      <div class="copy-row">
        <h3>Current Status</h3>
        <button class="copy-button" type="button" onclick="copyText('{status_id}', this)">Copy</button>
      </div>
      <pre id="{status_id}">{escape(json.dumps({"collector": status, "job": job, "progress": progress, "brandingRoute": route_status, "captureState": capture_state}, indent=2))}</pre>
      <div class="copy-row">
        <h3>Command Log</h3>
        <button class="copy-button" type="button" onclick="copyText('{log_id}', this)">Copy</button>
      </div>
      <pre id="{log_id}">{escape(log_text or "No command log yet.")}</pre>
    </section>
    """


def render_capture_manifest_summary(active_profile: str) -> str:
    manifest = capture_manifest()
    if not manifest:
        detail = '<p class="warn">No capture manifest loaded yet. After saving/uploading in Vercel admin, load it here so this tool captures the same Program IDs the public site expects.</p>'
    else:
        profiles = manifest.get("profiles") if isinstance(manifest.get("profiles"), dict) else {}
        rows = []
        for profile, config in PROFILES.items():
            row = profiles.get(profile, {}) if isinstance(profiles, dict) else {}
            ids = row.get("programIds") if isinstance(row, dict) else []
            count = len(ids) if isinstance(ids, list) else 0
            latest = row.get("latestSnapshotId") if isinstance(row, dict) else None
            rows.append(
                f"<span class=\"pill\">{escape(config['label'])}: {count} IDs; latest {escape(str(latest or 'none'))}</span>"
            )
        detail = (
            f"<p class=\"warn\"><strong>{escape(str(manifest.get('publicationTitle', 'Current publication')))}</strong> "
            f"({escape(str(manifest.get('publicationSlug', '')))}), generated "
            f"{escape(str(manifest.get('generatedAt', 'unknown')))}</p><div class=\"meta\">{''.join(rows)}</div>"
        )
    return f"""
    <section class="card" style="margin-bottom: 18px;">
      <h2>Vercel publication manifest</h2>
      {detail}
      <form method="post" action="{url_for('load_manifest', profile=active_profile)}">
        <div class="actions">
          <button type="submit" class="secondary">Load latest publication from Vercel</button>
        </div>
      </form>
    </section>
    """


@app.get("/")
def index() -> str:
    env = process_env()
    active_profile = selected_profile()
    cards = render_profile(active_profile, PROFILES[active_profile], env)
    token_status = blob_token_source(env)
    profile_tabs = "\n".join(
        f'<a class="{"active" if profile == active_profile else ""}" href="{url_for("index", profile=profile)}">{escape(config["label"])}</a>'
        for profile, config in PROFILES.items()
    )
    body = f"""
    <header>
      <div>
        <h1>CAS Branding<br>Capture</h1>
      </div>
      <p class="lede">
        Local control panel for WebAdMIT branding. It uses your browser login, captures each Program ID,
        uploads normalized branding JSON and images to Vercel Blob, and lets the deployed app read the latest snapshot.
      </p>
    </header>
    <nav class="profile-tabs" aria-label="CAS profile">
      {profile_tabs}
    </nav>
    <section class="card" style="margin-bottom: 18px;">
      <h2>Shared settings</h2>
      <div class="meta">
        <span class="pill">Blob token: {token_status}</span>
        <span><strong>WebAdMIT / CAS start URL:</strong> {escape(branding_start_url(env))}</span>
      </div>
      <form method="post" action="{url_for('save_start_url', profile=active_profile)}">
        <label for="start-url">WebAdMIT / CAS start URL</label>
        <input id="start-url" name="start_url" value="{escape(branding_start_url(env))}">
        <div class="actions">
          <button type="submit" class="secondary">Save URL</button>
        </div>
      </form>
      <form method="post" action="{url_for('save_blob_token', profile=active_profile)}">
        <label for="blob-token">BLOB_READ_WRITE_TOKEN</label>
        <input id="blob-token" name="blob_token" type="password" placeholder="Paste token from Vercel Blob store">
        <div class="actions">
          <button type="submit" class="secondary">Save token locally</button>
        </div>
      </form>
      <p class="warn">This saves to .env.local, which is ignored by git. The deployed Vercel app still needs the same token configured in Vercel.</p>
    </section>
    {render_capture_manifest_summary(active_profile)}
    <div class="grid">{cards}</div>
    """
    return page_shell(body)


@app.post("/settings/blob-token")
def save_blob_token():
    token = (request.form.get("blob_token") or "").strip()
    if token:
        save_env_value(REPO_ROOT / ".env.local", "BLOB_READ_WRITE_TOKEN", token)
    return profile_redirect(selected_profile())


@app.post("/settings/start-url")
def save_start_url():
    value = (request.form.get("start_url") or "").strip()
    if value:
        save_env_value(REPO_ROOT / ".env.local", "BRANDING_START_URL", value)
    return profile_redirect(selected_profile())


@app.post("/manifest/load")
def load_manifest():
    profile = selected_profile()
    thread = threading.Thread(target=load_capture_manifest, args=(profile,), daemon=True)
    thread.start()
    return profile_redirect(profile)


@app.post("/guide/<profile>")
def guide(profile: str):
    if profile not in PROFILES:
        return "Unknown profile", 404
    env = process_env()
    command = [
        "node",
        "tools/branding/cli.mjs",
        "guide",
        "--profile",
        profile,
        "--auth-file",
        str(profile_root(profile) / "user.json"),
        "--trail-file",
        str(profile_root(profile) / "trail.json"),
        "--status-file",
        str(status_path(profile)),
        "--login-url",
        branding_start_url(env),
        "--non-interactive",
    ]
    start_thread(profile, "guide", command)
    return profile_redirect(profile)


@app.post("/upload-export/<profile>")
def upload_export(profile: str):
    if profile not in PROFILES:
        return "Unknown profile", 404
    uploaded = request.files.get("xlsx_file")
    if not uploaded or not uploaded.filename:
        write_job(profile, status="error", message="No Excel file selected.", completedAt=utc_now())
        return profile_redirect(profile)
    filename = secure_filename(uploaded.filename)
    if not filename.lower().endswith((".xlsx", ".xls")):
        write_job(profile, status="error", message="Choose an .xlsx or .xls file.", completedAt=utc_now())
        return profile_redirect(profile)
    target_dir = UPLOAD_ROOT / profile
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / filename
    uploaded.save(target)
    write_job(
        profile,
        status="completed",
        action="upload_export",
        completedAt=utc_now(),
        message=f"Using Excel export: {target}",
    )
    return profile_redirect(profile)


@app.post("/capture/<profile>")
def capture(profile: str):
    if profile not in PROFILES:
        return "Unknown profile", 404
    env = process_env()
    route_status = capture_route_status(profile, env)
    manifest_ids = manifest_program_ids(profile)
    if not env.get("BLOB_READ_WRITE_TOKEN"):
        write_job(profile, status="error", message="Blob token is missing.", completedAt=utc_now())
        return profile_redirect(profile)
    if not route_status["ready"]:
        write_job(
            profile,
            status="error",
            message=f"{route_status['message']} Open guided login, reach Branding, then close Edge while Branding is visible.",
            completedAt=utc_now(),
        )
        return profile_redirect(profile)
    xlsx = Path(request.form.get("xlsx") or str(PROFILES[profile]["xlsx"])).expanduser()
    if not xlsx.is_absolute():
        xlsx = REPO_ROOT / xlsx
    if not manifest_ids and not xlsx.exists():
        write_job(
            profile,
            status="error",
            message="No Program IDs available. Load the Vercel manifest or choose an Excel report.",
            completedAt=utc_now(),
        )
        return profile_redirect(profile)
    try:
        delay_ms = int(request.form.get("delay_ms") or "4500")
        if delay_ms < 0 or delay_ms > 120_000:
            delay_ms = 4500
    except (TypeError, ValueError):
        delay_ms = 4500
    thread = threading.Thread(
        target=capture_and_upload,
        args=(profile, xlsx, delay_ms),
        daemon=True,
    )
    thread.start()
    return profile_redirect(profile)


@app.post("/upload-latest/<profile>")
def upload_latest(profile: str):
    if profile not in PROFILES:
        return "Unknown profile", 404
    thread = threading.Thread(target=upload_latest_snapshot, args=(profile,), daemon=True)
    thread.start()
    return profile_redirect(profile)


if __name__ == "__main__":
    print(f"Python: {sys.executable}")
    print("Open http://127.0.0.1:5050")
    app.run(host="127.0.0.1", port=5050, debug=False)
